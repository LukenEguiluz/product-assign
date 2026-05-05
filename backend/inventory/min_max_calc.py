from __future__ import annotations

import math
import re
import unicodedata
from datetime import datetime
from io import BytesIO
from typing import Any

import numpy as np
import pandas as pd
from openpyxl.utils import get_column_letter

LEAD_TIME_DIAS = 7
PERIODO_REABASTECIMIENTO_DIAS = 7
Z_NIVEL_SERVICIO = 1.65  # ~95 %

# Si no hay consumo en el periodo (resumen vs inventario): mínimo, máximo e ideal
DEFAULT_MIN_MAX_IDEAL_SIN_CONSUMO = 1


def _sin_acentos(s: str) -> str:
    s = unicodedata.normalize("NFKD", s)
    return "".join(c for c in s if not unicodedata.combining(c))


def _normalizar_nombre_col(nombre: str) -> str:
    s = _sin_acentos(str(nombre).strip().lower())
    s = re.sub(r"\s+", " ", s)
    return s


def _mapear_columnas(columns: list[str]) -> dict[str, str]:
    """normalizado -> nombre real en el DataFrame."""
    return {_normalizar_nombre_col(c): c for c in columns}


def detectar_columna_producto(norm_map: dict[str, str]) -> str | None:
    candidatos = [
        "codigo",
        "código",
        "producto",
        "sku",
        "clave",
        "articulo",
        "artículo",
        "id producto",
    ]
    for key in candidatos:
        k = _normalizar_nombre_col(key)
        if k in norm_map:
            return norm_map[k]
    for nk, real in norm_map.items():
        if "codigo" in nk or nk.startswith("sku"):
            return real
    return None


def detectar_columna_consumo(norm_map: dict[str, str]) -> str | None:
    candidatos = [
        "consumo",
        "cantidad",
        "qty",
        "unidades",
        "piezas",
        "demanda",
    ]
    for key in candidatos:
        k = _normalizar_nombre_col(key)
        if k in norm_map:
            return norm_map[k]
    for nk, real in norm_map.items():
        if any(x in nk for x in ("consumo", "cantidad", "unidades", "piezas")):
            return real
    return None


def detectar_columna_descripcion(norm_map: dict[str, str]) -> str | None:
    candidatos = [
        "descripcion",
        "descripción",
        "nombre",
        "nombre producto",
        "descripcion producto",
        "producto descripcion",
    ]
    for key in candidatos:
        k = _normalizar_nombre_col(key)
        if k in norm_map:
            return norm_map[k]
    for nk, real in norm_map.items():
        if nk.startswith("descripcion"):
            return real
    return None


def detectar_columna_documento(norm_map: dict[str, str]) -> str | None:
    candidatos = [
        "documento de reposicion",
        "documento de reposición",
        "documento reposicion",
        "documento reposición",
        "documento",
    ]
    for key in candidatos:
        k = _normalizar_nombre_col(key)
        if k in norm_map:
            return norm_map[k]
    for nk, real in norm_map.items():
        if nk.startswith("documento"):
            return real
    return None


def detectar_columna_fecha(norm_map: dict[str, str]) -> str | None:
    preferidas = [
        "fecha de salida",
        "fecha salida",
        "fecha consumo",
        "fecha movimiento",
        "fecha",
        "fecha de uso",
    ]
    for key in preferidas:
        k = _normalizar_nombre_col(key)
        if k in norm_map:
            return norm_map[k]
    for nk, real in norm_map.items():
        if nk.startswith("fecha") or " fecha" in nk:
            return real
    return None


def _descripcion_modal_en_grupo(grupo: pd.DataFrame, col_descripcion: str | None) -> str:
    if not col_descripcion or col_descripcion not in grupo.columns:
        return ""
    vals = grupo[col_descripcion].dropna().astype(str).str.strip()
    vals = vals[vals != ""]
    if vals.empty:
        return ""
    mode = vals.mode()
    return str(mode.iloc[0]) if len(mode) else str(vals.iloc[0])


def _valor_modal_en_grupo(grupo: pd.DataFrame, col: str | None) -> str:
    if not col or col not in grupo.columns:
        return ""
    vals = grupo[col].dropna().astype(str).str.strip()
    vals = vals[vals != ""]
    if vals.empty:
        return ""
    mode = vals.mode()
    return str(mode.iloc[0]) if len(mode) else str(vals.iloc[0])


def agregar_inventario_por_referencia(
    df_inv: pd.DataFrame,
    col_producto: str,
    col_descripcion: str | None,
    col_documento: str | None,
) -> pd.DataFrame:
    """Stock actual = número de filas por referencia (misma lógica que conteo de consumos)."""
    filas: list[dict[str, Any]] = []
    for ref, grupo in df_inv.groupby(col_producto, dropna=False):
        filas.append(
            {
                "Referencia": ref,
                "Descripción": _descripcion_modal_en_grupo(grupo, col_descripcion),
                "Documento inventario": _valor_modal_en_grupo(grupo, col_documento),
                "Stock actual": int(len(grupo)),
            }
        )
    out = pd.DataFrame(filas)
    out["Referencia"] = out["Referencia"].astype(str).str.strip()
    return out


def estado_vs_ideal(actual: int, ideal: int) -> str:
    if actual < ideal:
        return "Por debajo del ideal"
    if actual > ideal:
        return "Exceso vs ideal"
    return "En ideal"


def construir_resumen_inventario_vs_calculo(
    df_inv_agg: pd.DataFrame,
    df_calculo: pd.DataFrame,
    col_ref_calculo: str,
) -> pd.DataFrame:
    """df_calculo debe incluir col_ref_calculo, Stock mínimo recomendado, Stock máximo recomendado."""
    calc = df_calculo[
        [
            col_ref_calculo,
            "Documento consumo",
            "Stock mínimo recomendado",
            "Stock máximo recomendado",
        ]
    ].copy()
    calc["_key"] = calc[col_ref_calculo].astype(str).str.strip()
    calc = calc.drop_duplicates(subset="_key", keep="first")

    res = df_inv_agg.merge(
        calc,
        left_on="Referencia",
        right_on="_key",
        how="left",
    )
    drop_cols = ["_key"]
    if col_ref_calculo in res.columns:
        drop_cols.append(col_ref_calculo)
    res = res.drop(columns=[c for c in drop_cols if c in res.columns])

    d = DEFAULT_MIN_MAX_IDEAL_SIN_CONSUMO
    res["Documento consumo"] = res["Documento consumo"].fillna("")
    res["Stock mínimo recomendado"] = (
        pd.to_numeric(res["Stock mínimo recomendado"], errors="coerce")
        .fillna(d)
        .astype(int)
    )
    res["Stock máximo recomendado"] = (
        pd.to_numeric(res["Stock máximo recomendado"], errors="coerce")
        .fillna(d)
        .astype(int)
    )
    res["Stock ideal"] = (
        ((res["Stock mínimo recomendado"] + res["Stock máximo recomendado"]) / 2.0)
        .round()
        .astype(int)
    )
    res["Estado"] = [
        estado_vs_ideal(int(a), int(i))
        for a, i in zip(res["Stock actual"], res["Stock ideal"])
    ]

    res = res.rename(
        columns={
            "Stock mínimo recomendado": "Stock mínimo",
            "Stock máximo recomendado": "Stock máximo",
        }
    )

    columnas = [
        "Referencia",
        "Descripción",
        "Documento consumo",
        "Documento inventario",
        "Stock ideal",
        "Stock mínimo",
        "Stock máximo",
        "Stock actual",
        "Estado",
    ]
    for c in ["Stock ideal", "Stock mínimo", "Stock máximo"]:
        res[c] = pd.to_numeric(res[c], errors="coerce").astype("Int64")
    res["Stock actual"] = res["Stock actual"].astype("Int64")
    return res[columnas]


def _read_excel_from_bytes(xlsx_bytes: bytes, sheet_name: str | int | None) -> pd.DataFrame:
    bio = BytesIO(xlsx_bytes)
    hoja: str | int = 0 if sheet_name in (None, "") else sheet_name
    df = pd.read_excel(bio, sheet_name=hoja)
    if df is None or df.empty:
        raise ValueError("El archivo no tiene filas de datos.")
    return df


def build_comparativa_min_max_xlsx(
    *,
    consumo_xlsx: bytes,
    inventario_xlsx: bytes,
    hoja_consumo: str | int | None = None,
    hoja_inventario: str | int | None = None,
    meses: int = 6,
    fecha_referencia: str | None = None,  # YYYY-MM-DD
    lead_time_dias: int = LEAD_TIME_DIAS,
    periodo_reabastecimiento_dias: int = PERIODO_REABASTECIMIENTO_DIAS,
    z_score: float = Z_NIVEL_SERVICIO,
    col_producto: str | None = None,
    col_consumo: str | None = None,
    col_fecha: str | None = None,
    col_descripcion: str | None = None,
    col_documento: str | None = None,
    cliente_nombre: str | None = None,
    cliente_numero: str | None = None,
    fecha_generacion_documento: str | None = None,  # YYYY-MM-DD
) -> bytes:
    df = _read_excel_from_bytes(consumo_xlsx, hoja_consumo)
    norm_map = _mapear_columnas(list(df.columns))

    col_producto = col_producto or detectar_columna_producto(norm_map)
    if not col_producto:
        raise ValueError(
            "No se detectó columna de producto en consumos. Indica una columna (col_producto)."
        )

    col_descripcion = col_descripcion or detectar_columna_descripcion(norm_map)
    col_documento = col_documento or detectar_columna_documento(norm_map)

    col_fecha = col_fecha or detectar_columna_fecha(norm_map)
    if not col_fecha:
        raise ValueError(
            "No se detectó columna de fecha en consumos. Indica una columna (col_fecha)."
        )

    col_consumo = col_consumo or detectar_columna_consumo(norm_map)
    modo_conteo = col_consumo is None
    if not modo_conteo:
        df["_consumo_fila"] = pd.to_numeric(df[col_consumo], errors="coerce").fillna(0)
    else:
        df["_consumo_fila"] = 1.0

    fechas = pd.to_datetime(df[col_fecha], errors="coerce")
    df = df.assign(_fecha=fechas).dropna(subset=["_fecha"])

    if fecha_referencia:
        fin = pd.Timestamp(datetime.strptime(fecha_referencia, "%Y-%m-%d")).normalize()
    else:
        fin = pd.Timestamp.now().normalize()

    inicio = (fin - pd.DateOffset(months=meses)).normalize()
    df["_dia"] = df["_fecha"].dt.normalize()
    df = df[(df["_dia"] >= inicio) & (df["_dia"] <= fin)]
    if df.empty:
        raise ValueError(
            f"No hay registros entre {inicio.date()} y {fin.date()} (últimos {meses} meses)."
        )

    dias_calendario = pd.date_range(inicio.normalize(), fin.normalize(), freq="D")
    n_dias = len(dias_calendario)

    def r2(x: float) -> float:
        return round(float(x), 2)

    def stock_ceil_2(x: float) -> float:
        xf = float(x)
        return math.ceil(round(xf, 12) * 100) / 100

    def stock_ceil_int(x: float) -> int:
        return int(math.ceil(round(float(x), 12)))

    def descripcion_por_producto(grupo: pd.DataFrame) -> str:
        if not col_descripcion or col_descripcion not in grupo.columns:
            return ""
        vals = grupo[col_descripcion].dropna().astype(str).str.strip()
        vals = vals[vals != ""]
        if vals.empty:
            return ""
        mode = vals.mode()
        return str(mode.iloc[0]) if len(mode) else str(vals.iloc[0])

    resultados: list[dict[str, Any]] = []
    for producto, grupo in df.groupby(col_producto, dropna=False):
        diario = grupo.groupby("_dia", as_index=True)["_consumo_fila"].sum()
        serie = diario.reindex(dias_calendario, fill_value=0.0)

        demanda_promedio = float(serie.mean())
        desviacion = float(serie.std(ddof=1)) if n_dias > 1 else 0.0
        if np.isnan(desviacion):
            desviacion = 0.0

        stock_seguridad = float(z_score) * desviacion * np.sqrt(float(lead_time_dias))
        stock_min = (demanda_promedio * float(lead_time_dias)) + stock_seguridad
        stock_max = stock_min + (demanda_promedio * float(periodo_reabastecimiento_dias))

        fila: dict[str, Any] = {
            col_producto: producto,
            "Demanda promedio diaria": r2(demanda_promedio),
            "Desviación demanda diaria": r2(desviacion),
            "Stock de seguridad": stock_ceil_2(stock_seguridad),
            "Total consumo periodo": r2(float(serie.sum())),
            "Días con movimiento": int((serie > 0).sum()),
            "Días calendario periodo": int(n_dias),
            "Stock mínimo recomendado": stock_ceil_int(stock_min),
            "Stock máximo recomendado": stock_ceil_int(stock_max),
        }
        if col_descripcion:
            fila[col_descripcion] = descripcion_por_producto(grupo)
        fila["Documento consumo"] = _valor_modal_en_grupo(grupo, col_documento)
        resultados.append(fila)

    df_resultado = pd.DataFrame(resultados)

    columnas_salida: list[str] = [col_producto]
    if col_descripcion:
        columnas_salida.append(col_descripcion)
    columnas_salida.append("Documento consumo")
    columnas_salida.extend(
        [
            "Demanda promedio diaria",
            "Desviación demanda diaria",
            "Stock de seguridad",
            "Total consumo periodo",
            "Días con movimiento",
            "Días calendario periodo",
            "Stock mínimo recomendado",
            "Stock máximo recomendado",
        ]
    )
    df_resultado = df_resultado[columnas_salida]

    columnas_redondeo_normal = [
        "Demanda promedio diaria",
        "Desviación demanda diaria",
        "Total consumo periodo",
    ]
    for col in columnas_redondeo_normal:
        df_resultado[col] = (
            pd.to_numeric(df_resultado[col], errors="coerce").astype(float).round(2)
        )

    columnas_stock_seguridad_ceil2 = ["Stock de seguridad"]
    for col in columnas_stock_seguridad_ceil2:
        df_resultado[col] = (
            pd.to_numeric(df_resultado[col], errors="coerce")
            .astype(float)
            .map(lambda v: stock_ceil_2(v) if pd.notna(v) else v)
        )

    columnas_stock_min_max_entero = [
        "Stock mínimo recomendado",
        "Stock máximo recomendado",
    ]
    for col in columnas_stock_min_max_entero:
        df_resultado[col] = (
            pd.to_numeric(df_resultado[col], errors="coerce")
            .map(lambda v: stock_ceil_int(v) if pd.notna(v) else v)
            .astype("Int64")
        )

    df_inv = _read_excel_from_bytes(inventario_xlsx, hoja_inventario)
    norm_inv = _mapear_columnas(list(df_inv.columns))
    col_prod_inv = detectar_columna_producto(norm_inv) or col_producto
    if not col_prod_inv:
        raise ValueError(
            "No se detectó columna de producto en inventario. Indica una columna (col_producto)."
        )
    col_desc_inv = detectar_columna_descripcion(norm_inv) or col_descripcion
    col_doc_inv = detectar_columna_documento(norm_inv)

    df_inv_agg = agregar_inventario_por_referencia(
        df_inv, col_prod_inv, col_desc_inv, col_doc_inv
    )
    df_resumen = construir_resumen_inventario_vs_calculo(
        df_inv_agg, df_resultado, col_producto
    )

    meta_dict: dict[str, Any] = {
        "hoja_consumo": str(0 if hoja_consumo in (None, "") else hoja_consumo),
        "hoja_inventario": str(0 if hoja_inventario in (None, "") else hoja_inventario),
        "periodo_inicio": inicio.date().isoformat(),
        "periodo_fin": fin.date().isoformat(),
        "meses": int(meses),
        "col_producto": col_producto,
        "col_descripcion": col_descripcion or "",
        "col_documento": col_documento or "",
        "col_fecha": col_fecha,
        "col_consumo": col_consumo if not modo_conteo else "(conteo por fila = 1)",
        "lead_time_dias": int(lead_time_dias),
        "periodo_reabastecimiento_dias": int(periodo_reabastecimiento_dias),
        "z": round(float(z_score), 2),
        "stock_ideal": "Entero: redondeo de (mín + máx) / 2",
        "sin_consumo_min_max_ideal": (
            f"{DEFAULT_MIN_MAX_IDEAL_SIN_CONSUMO} si no hay registro de consumo en el periodo"
        ),
        "stock_actual": "Entero: filas por referencia en inventario",
    }
    if cliente_nombre:
        meta_dict["cliente"] = str(cliente_nombre).strip()
    if cliente_numero is not None:
        meta_dict["numero_cliente"] = str(cliente_numero).strip()
    if fecha_generacion_documento:
        meta_dict["fecha_generacion_documento"] = fecha_generacion_documento
    meta = pd.DataFrame([meta_dict])

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df_resumen.to_excel(writer, sheet_name="Resumen", index=False)
        meta.to_excel(writer, sheet_name="Parametros", index=False)

        ws = writer.sheets["Resumen"]
        header_a_col: dict[str, int] = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=1, column=c).value
            if v is not None:
                header_a_col[str(v)] = c
        for nombre in (
            "Stock ideal",
            "Stock mínimo",
            "Stock máximo",
            "Stock actual",
        ):
            idx = header_a_col.get(nombre)
            if idx is None:
                continue
            letter = get_column_letter(idx)
            for r in range(2, ws.max_row + 1):
                ws[f"{letter}{r}"].number_format = "0"

    return buf.getvalue()

