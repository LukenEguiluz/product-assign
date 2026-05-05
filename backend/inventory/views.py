from django.contrib.auth.models import User
from django.db.models import Count, Prefetch, Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from django.utils.text import slugify
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from rest_framework import generics, status, viewsets
from rest_framework.decorators import action
from rest_framework.pagination import PageNumberPagination
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework_simplejwt.views import TokenObtainPairView

from .catalog_import import build_template_workbook, parse_catalog_workbook
from .min_max_calc import build_comparativa_min_max_xlsx
from .models import Cabinet, CatalogItem, Client, InventorySession, SessionScan
from .permissions import CanCreateAppUsers, IsSuperuser
from .serializers import (
    AddScanSerializer,
    CabinetSerializer,
    CatalogItemSerializer,
    ClientSerializer,
    CreateAppUserSerializer,
    InventorySessionDetailSerializer,
    InventorySessionSerializer,
    SessionScanSerializer,
    SessionScanUpdateSerializer,
    SetAppUserPasswordSerializer,
    UserSerializer,
)


class LoginView(TokenObtainPairView):
    permission_classes = [AllowAny]


class MeView(APIView):
    def get(self, request):
        data = UserSerializer(request.user).data
        data["can_create_app_users"] = request.user.has_perm(
            "inventory.can_create_app_users"
        )
        data["is_superuser"] = bool(request.user.is_superuser)
        return Response(data)


class CreateAppUserView(generics.CreateAPIView):
    serializer_class = CreateAppUserSerializer
    permission_classes = [IsAuthenticated, CanCreateAppUsers]


class UserListView(generics.ListAPIView):
    queryset = User.objects.order_by("username")
    serializer_class = UserSerializer
    permission_classes = [IsAuthenticated, CanCreateAppUsers]


class SetAppUserPasswordView(APIView):
    permission_classes = [IsAuthenticated, IsSuperuser]

    def post(self, request, user_id: int):
        target = get_object_or_404(User.objects.all(), pk=user_id)
        ser = SetAppUserPasswordSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        target.set_password(ser.validated_data["password"])
        target.save(update_fields=["password"])
        return Response(status=status.HTTP_204_NO_CONTENT)


class ClientViewSet(viewsets.ModelViewSet):
    queryset = Client.objects.all()
    serializer_class = ClientSerializer


class CabinetViewSet(viewsets.ModelViewSet):
    queryset = Cabinet.objects.select_related("client").all()
    serializer_class = CabinetSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        client_id = self.request.query_params.get("client")
        if client_id:
            qs = qs.filter(client_id=client_id)
        return qs


class CatalogItemPagination(PageNumberPagination):
    page_size = 50
    page_size_query_param = "page_size"
    max_page_size = 200


class CatalogItemViewSet(viewsets.ModelViewSet):
    queryset = CatalogItem.objects.all().order_by("reference", "gtin")
    serializer_class = CatalogItemSerializer
    lookup_field = "gtin"
    pagination_class = CatalogItemPagination

    def get_queryset(self):
        qs = super().get_queryset()
        if self.action != "list":
            return qs
        raw = (
            self.request.query_params.get("q")
            or self.request.query_params.get("search")
            or ""
        ).strip()
        if raw:
            qs = qs.filter(
                Q(gtin__icontains=raw)
                | Q(reference__icontains=raw)
                | Q(description__icontains=raw)
            )
        return qs

    def get_object(self):
        raw = self.kwargs.get("gtin", "") or ""
        self.kwargs["gtin"] = raw.strip().upper()
        return super().get_object()

    @action(detail=False, methods=["get"], url_path="import-template")
    def import_template(self, request):
        buf = build_template_workbook()
        response = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            'attachment; filename="plantilla_catalogo.xlsx"'
        )
        return response

    @action(detail=False, methods=["post"], url_path="import-excel")
    def import_excel(self, request):
        upload = request.FILES.get("file")
        if not upload:
            return Response(
                {"detail": "Adjunte el archivo en el campo multipart «file»."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if not upload.name.lower().endswith(".xlsx"):
            return Response(
                {"detail": "Solo se admite formato .xlsx."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        raw = b"".join(upload.chunks())
        flag = request.data.get("update_existing")
        update_existing = str(flag or "").lower() in ("1", "true", "yes", "on")
        result = parse_catalog_workbook(raw, update_existing=update_existing)
        if result.get("error"):
            return Response(result, status=status.HTTP_400_BAD_REQUEST)
        return Response(result, status=status.HTTP_200_OK)


class InventorySessionViewSet(viewsets.ModelViewSet):
    queryset = InventorySession.objects.select_related(
        "client", "cabinet", "created_by"
    ).prefetch_related(
        Prefetch(
            "scans",
            queryset=SessionScan.objects.select_related(
                "created_by", "excluded_by"
            ).order_by("-created_at"),
        ),
    )
    serializer_class = InventorySessionSerializer

    def get_serializer_class(self):
        if self.action == "retrieve":
            return InventorySessionDetailSerializer
        return InventorySessionSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        qs = qs.annotate(
            _active_scan_count=Count(
                "scans",
                filter=Q(scans__excluded_at__isnull=True),
            ),
        )
        st = self.request.query_params.get("status")
        if st:
            qs = qs.filter(status=st)
        mine = self.request.query_params.get("mine")
        if mine in ("1", "true", "yes"):
            qs = qs.filter(created_by=self.request.user)
        return qs

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user, status=InventorySession.Status.DRAFT)

    @action(detail=True, methods=["post"], url_path="start")
    def start(self, request, pk=None):
        session = self.get_object()
        if session.status == InventorySession.Status.COMPLETED:
            return Response(
                {"detail": "La sesión ya está finalizada."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        session.status = InventorySession.Status.IN_PROGRESS
        session.save(update_fields=["status", "updated_at"])
        return Response(InventorySessionDetailSerializer(session).data)

    @action(detail=True, methods=["post"], url_path="pause")
    def pause(self, request, pk=None):
        session = self.get_object()
        if session.status == InventorySession.Status.COMPLETED:
            return Response(
                {"detail": "La sesión ya está finalizada."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        session.status = InventorySession.Status.DRAFT
        session.save(update_fields=["status", "updated_at"])
        return Response(InventorySessionDetailSerializer(session).data)

    @action(detail=True, methods=["post"], url_path="complete")
    def complete(self, request, pk=None):
        session = self.get_object()
        session.status = InventorySession.Status.COMPLETED
        session.completed_at = timezone.now()
        session.save(update_fields=["status", "completed_at", "updated_at"])
        return Response(InventorySessionDetailSerializer(session).data)

    @action(detail=True, methods=["post"], url_path="resume")
    def resume(self, request, pk=None):
        session = self.get_object()
        if session.status != InventorySession.Status.COMPLETED:
            return Response(
                {"detail": "Solo las lecturas finalizadas pueden reabrirse para continuar."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        session.status = InventorySession.Status.DRAFT
        session.completed_at = None
        session.save(update_fields=["status", "completed_at", "updated_at"])
        return Response(InventorySessionDetailSerializer(session).data)

    @action(detail=True, methods=["post"], url_path="scans")
    def add_scan(self, request, pk=None):
        session = self.get_object()
        if session.status == InventorySession.Status.COMPLETED:
            return Response(
                {"detail": "No se pueden agregar escaneos a una sesión finalizada."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if session.status == InventorySession.Status.DRAFT:
            session.status = InventorySession.Status.IN_PROGRESS
            session.save(update_fields=["status", "updated_at"])
        ser = AddScanSerializer(
            data=request.data,
            context={"session": session, "request": request},
        )
        ser.is_valid(raise_exception=True)
        scan = ser.save()
        return Response(
            SessionScanSerializer(scan).data, status=status.HTTP_201_CREATED
        )

    @action(detail=True, methods=["get"], url_path="export-excel")
    def export_excel(self, request, pk=None):
        """Exporta líneas activas en hoja «Inventario» con columnas fijas y tabla Excel."""
        session = self.get_object()

        scans = list(
            session.scans.filter(excluded_at__isnull=True).order_by("-created_at")
        )
        gtins = {s.gtin for s in scans if s.gtin}
        ref_by_gtin = {}
        if gtins:
            for ci in CatalogItem.objects.filter(gtin__in=gtins).only(
                "gtin", "reference"
            ):
                ref_by_gtin[ci.gtin] = ci.reference

        dlv = (session.delivery_number or "").strip()
        inv_date = session.inventory_date
        client_no = (session.client.client_number or "").strip()

        wb = Workbook()
        ws = wb.active
        ws.title = "Inventario"

        headers = [
            "Código",
            "Lote",
            "Caducidad",
            "Documento de reposición",
            "Fecha de reposición",
            "No. de envío",
            "Orden de compra",
            "Ticket de salida",
            "Almacén BSCI",
            "No. de cliente",
            "Etiqueta RFID",
        ]
        ws.append(headers)

        for s in scans:
            ref = (ref_by_gtin.get(s.gtin) or "").strip()
            ws.append(
                [
                    ref,
                    (s.batch_lot or "").strip(),
                    (s.expiry_yymmdd or "").strip(),
                    dlv,
                    inv_date,
                    dlv,
                    "S/O",
                    "S/T",
                    "CEDIS",
                    client_no,
                    s.rfid_hex,
                ]
            )

        last_row = ws.max_row
        last_col_letter = get_column_letter(len(headers))
        if last_row >= 2:
            tab = Table(
                displayName="InventarioLectura",
                ref=f"A1:{last_col_letter}{last_row}",
            )
            tab.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            ws.add_table(tab)

        base_name = slugify(f"{session.client.name} {inv_date}")
        if not base_name:
            base_name = slugify(f"cliente-{session.client_id}-{inv_date}") or f"lectura-{session.id}"
        fname = f"{base_name}.xlsx"
        if len(fname) > 200:
            fname = f"{base_name[:160]}.xlsx"

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{fname}"'
        wb.save(response)
        return response


class SessionScanDetailView(generics.RetrieveUpdateDestroyAPIView):
    """Detalle, edición (GTIN/RFID) y exclusión lógica de una línea de lectura."""

    permission_classes = [IsAuthenticated]
    lookup_url_kwarg = "pk"

    def get_queryset(self):
        return SessionScan.objects.filter(
            session_id=self.kwargs["session_pk"],
            session__created_by=self.request.user,
        )

    def get_serializer_class(self):
        if self.request.method in ("PATCH", "PUT"):
            return SessionScanUpdateSerializer
        return SessionScanSerializer

    def partial_update(self, request, *args, **kwargs):
        return self._write_scan(request, partial=True)

    def update(self, request, *args, **kwargs):
        return self._write_scan(request, partial=False)

    def _write_scan(self, request, partial):
        instance = self.get_object()
        ser = SessionScanUpdateSerializer(
            instance, data=request.data, partial=partial
        )
        ser.is_valid(raise_exception=True)
        ser.save()
        instance.refresh_from_db()
        return Response(SessionScanSerializer(instance).data)

    def perform_destroy(self, instance):
        if instance.excluded_at is None:
            instance.excluded_at = timezone.now()
            instance.excluded_by = self.request.user
            instance.save(update_fields=["excluded_at", "excluded_by"])


class SessionScanRestoreView(APIView):
    """Vuelve a incluir en el inventario final una línea previamente excluida."""

    permission_classes = [IsAuthenticated]

    def post(self, request, session_pk, pk):
        scan = get_object_or_404(
            SessionScan.objects.filter(
                session_id=session_pk,
                session__created_by=request.user,
            ),
            pk=pk,
        )
        if scan.excluded_at is None:
            return Response(
                {"detail": "Este registro ya está incluido en el inventario final."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        conflict = SessionScan.objects.filter(
            session_id=session_pk,
            rfid_hex=scan.rfid_hex,
            excluded_at__isnull=True,
        ).exclude(pk=scan.pk).exists()
        if conflict:
            return Response(
                {
                    "detail": "Ya existe otro registro activo con el mismo RFID. Corrija o excluya el duplicado antes de restaurar."
                },
                status=status.HTTP_400_BAD_REQUEST,
            )
        scan.excluded_at = None
        scan.excluded_by = None
        scan.save(update_fields=["excluded_at", "excluded_by"])
        return Response(SessionScanSerializer(scan).data)


class CatalogLookupView(APIView):
    def get(self, request, gtin: str):
        code = (gtin or "").strip().upper()
        try:
            item = CatalogItem.objects.get(gtin=code)
        except CatalogItem.DoesNotExist:
            return Response({"exists": False, "gtin": code})
        return Response(
            {
                "exists": True,
                "item": CatalogItemSerializer(item).data,
            }
        )


class MinMaxStockView(APIView):
    """
    POST multipart/form-data
    - consumo: archivo .xlsx de consumos
    - inventario: archivo .xlsx de inventario
    - client: id del cliente registrado (para nombre del archivo y parámetros del documento)

    Opcionales (form fields):
    - meses (int), fecha_referencia (YYYY-MM-DD), lead_time_dias, periodo_reabastecimiento_dias, z_score
    - hoja_consumo, hoja_inventario (nombre o índice)
    - col_producto, col_consumo, col_fecha, col_descripcion, col_documento
    """

    permission_classes = [IsAuthenticated]

    def post(self, request):
        consumo = request.FILES.get("consumo")
        inventario = request.FILES.get("inventario")
        client_raw = request.data.get("client")
        if client_raw in (None, ""):
            return Response(
                {"detail": "Indique el cliente dirigido (campo «client»: id numérico del cliente)."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        try:
            client_pk = int(str(client_raw))
        except (TypeError, ValueError):
            return Response(
                {"detail": "El campo «client» debe ser un identificador numérico válido."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        client_obj = get_object_or_404(Client, pk=client_pk)

        gen_date = timezone.localdate()
        base_name = slugify(f"{client_obj.name} {gen_date.isoformat()}") or ""
        if not base_name:
            base_name = slugify(f"cliente-{client_obj.pk}-{gen_date.isoformat()}") or (
                f"cliente-{client_obj.pk}-min-max-{gen_date.isoformat()}"
            )

        if not consumo or not inventario:
            return Response(
                {
                    "detail": "Adjunte ambos archivos en multipart: «consumo» e «inventario».",
                },
                status=status.HTTP_400_BAD_REQUEST,
            )
        for f in (consumo, inventario):
            if not (f.name or "").lower().endswith(".xlsx"):
                return Response(
                    {"detail": "Solo se admite formato .xlsx."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        consumo_bytes = b"".join(consumo.chunks())
        inventario_bytes = b"".join(inventario.chunks())

        def _as_int(key: str, default: int):
            raw = request.data.get(key, None)
            if raw in (None, ""):
                return default
            try:
                return int(str(raw))
            except (TypeError, ValueError):
                raise ValueError(f"Parámetro inválido: {key}")

        def _as_float(key: str, default: float):
            raw = request.data.get(key, None)
            if raw in (None, ""):
                return default
            try:
                return float(str(raw))
            except (TypeError, ValueError):
                raise ValueError(f"Parámetro inválido: {key}")

        def _as_sheet(key: str):
            raw = request.data.get(key, None)
            if raw in (None, ""):
                return None
            s = str(raw).strip()
            if s.isdigit():
                return int(s)
            return s

        try:
            out = build_comparativa_min_max_xlsx(
                consumo_xlsx=consumo_bytes,
                inventario_xlsx=inventario_bytes,
                hoja_consumo=_as_sheet("hoja_consumo"),
                hoja_inventario=_as_sheet("hoja_inventario"),
                meses=_as_int("meses", 6),
                fecha_referencia=(request.data.get("fecha_referencia") or None),
                lead_time_dias=_as_int("lead_time_dias", 7),
                periodo_reabastecimiento_dias=_as_int(
                    "periodo_reabastecimiento_dias", 7
                ),
                z_score=_as_float("z_score", 1.65),
                col_producto=(request.data.get("col_producto") or None),
                col_consumo=(request.data.get("col_consumo") or None),
                col_fecha=(request.data.get("col_fecha") or None),
                col_descripcion=(request.data.get("col_descripcion") or None),
                col_documento=(request.data.get("col_documento") or None),
                cliente_nombre=client_obj.name,
                cliente_numero=client_obj.client_number,
                fecha_generacion_documento=gen_date.isoformat(),
            )
        except ValueError as e:
            return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception:
            return Response(
                {"detail": "No se pudo procesar los archivos. Revise el formato/columnas."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        fname = f"{base_name}.xlsx"
        if len(fname) > 200:
            fname = f"{base_name[:160]}.xlsx"

        response = HttpResponse(
            out,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{fname}"'
        return response
