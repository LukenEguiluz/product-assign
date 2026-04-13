"""Importación masiva de catálogo desde Excel (.xlsx)."""
from __future__ import annotations

import re
from collections import defaultdict
from io import BytesIO
from typing import Any

from django.db import transaction
from django.utils import timezone
from openpyxl import Workbook, load_workbook

from .models import CatalogItem

GTIN_ALIASES = frozenset({"GTIN", "GUDID", "EAN", "CODIGO", "CÓDIGO", "CODE"})
REF_ALIASES = frozenset({"REFERENCIA", "REF", "REFERENCE", "SKU"})
DESC_ALIASES = frozenset({"DESCRIPCION", "DESCRIPCIÓN", "DESC", "DESCRIPTION", "DESCR"})


def _norm_header(cell: Any) -> str:
    if cell is None:
        return ""
    return re.sub(r"\s+", " ", str(cell).strip().upper())


def _norm_cell(val: Any) -> str:
    if val is None:
        return ""
    return str(val).strip().upper()


def _find_columns(header_row: list[Any]) -> tuple[int | None, int | None, int | None]:
    gtin_i = ref_i = desc_i = None
    for idx, cell in enumerate(header_row):
        h = _norm_header(cell)
        if h in GTIN_ALIASES:
            gtin_i = idx
        elif h in REF_ALIASES:
            ref_i = idx
        elif h in DESC_ALIASES:
            desc_i = idx
    return gtin_i, ref_i, desc_i


def _empty_result(error: str | None = None) -> dict[str, Any]:
    return {
        "error": error,
        "created_count": 0,
        "updated_count": 0,
        "created": [],
        "updated": [],
        "duplicates_in_file": [],
        "already_in_catalog": [],
        "invalid_rows": [],
        "skipped_duplicate_in_file_rows": 0,
        "update_existing": False,
    }


def parse_catalog_workbook(
    file_bytes: bytes,
    *,
    update_existing: bool,
) -> dict[str, Any]:
    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if not header:
            return _empty_result("El archivo no tiene filas.")

        gtin_i, ref_i, desc_i = _find_columns(list(header))
        if gtin_i is None or ref_i is None:
            return _empty_result(
                "Falta columna GTIN (o GUDID/EAN) y/o REFERENCIA en la primera fila."
            )

        parsed: list[dict[str, Any]] = []
        invalid_rows: list[dict[str, Any]] = []
        row_num = 1

        for row in rows_iter:
            row_num += 1
            if not row:
                continue
            cells = list(row)

            def cell(i: int | None) -> str:
                if i is None or i >= len(cells):
                    return ""
                return _norm_cell(cells[i])

            gtin = cell(gtin_i)
            reference = cell(ref_i)
            description = cell(desc_i) if desc_i is not None else ""

            if not gtin and not reference and not description:
                continue
            if not gtin:
                invalid_rows.append(
                    {"row": row_num, "reason": "GTIN vacío en fila con otros datos."}
                )
                continue
            if not reference:
                invalid_rows.append(
                    {
                        "row": row_num,
                        "reason": "REFERENCIA vacía (obligatoria junto con GTIN).",
                    }
                )
                continue
            if len(gtin) > 32:
                invalid_rows.append(
                    {"row": row_num, "reason": f"GTIN demasiado largo ({len(gtin)} > 32)."}
                )
                continue
            if len(reference) > 255:
                invalid_rows.append(
                    {"row": row_num, "reason": "REFERENCIA demasiado larga (> 255)."}
                )
                continue

            parsed.append(
                {
                    "row": row_num,
                    "gtin": gtin,
                    "reference": reference,
                    "description": description,
                }
            )

        by_gtin: dict[str, list[int]] = defaultdict(list)
        for p in parsed:
            by_gtin[p["gtin"]].append(p["row"])

        duplicates_in_file = [
            {"gtin": g, "rows": sorted(rows)}
            for g, rows in by_gtin.items()
            if len(rows) > 1
        ]
        dup_gtins = {g for g, rows in by_gtin.items() if len(rows) > 1}

        skipped_dup_file = 0
        already_in_catalog: list[dict[str, Any]] = []
        to_create: list[CatalogItem] = []
        to_update: list[CatalogItem] = []

        candidate_gtins = {p["gtin"] for p in parsed if p["gtin"] not in dup_gtins}
        existing_objs: dict[str, CatalogItem] = {}
        if candidate_gtins:
            existing_objs = {
                obj.gtin: obj
                for obj in CatalogItem.objects.filter(gtin__in=candidate_gtins)
            }

        now = timezone.now()

        for p in parsed:
            g = p["gtin"]
            if g in dup_gtins:
                skipped_dup_file += 1
                continue
            ex = existing_objs.get(g)
            if ex is not None:
                if update_existing:
                    ex.reference = p["reference"]
                    ex.description = p["description"]
                    ex.updated_at = now
                    if ex not in to_update:
                        to_update.append(ex)
                else:
                    already_in_catalog.append(
                        {
                            "gtin": g,
                            "row": p["row"],
                            "existing_reference": ex.reference,
                            "existing_description": ex.description,
                        }
                    )
                continue
            to_create.append(
                CatalogItem(
                    gtin=g,
                    reference=p["reference"],
                    description=p["description"],
                )
            )

        created: list[dict[str, str]] = []
        updated: list[dict[str, str]] = []

        with transaction.atomic():
            if to_create:
                CatalogItem.objects.bulk_create(to_create)
                created = [
                    {"gtin": obj.gtin, "reference": obj.reference} for obj in to_create
                ]
            if to_update:
                CatalogItem.objects.bulk_update(
                    to_update, ["reference", "description", "updated_at"]
                )
                updated = [
                    {"gtin": obj.gtin, "reference": obj.reference} for obj in to_update
                ]

        return {
            "error": None,
            "created_count": len(created),
            "updated_count": len(updated),
            "created": created,
            "updated": updated,
            "duplicates_in_file": duplicates_in_file,
            "already_in_catalog": already_in_catalog,
            "invalid_rows": invalid_rows,
            "skipped_duplicate_in_file_rows": skipped_dup_file,
            "update_existing": update_existing,
        }
    finally:
        wb.close()


def build_template_workbook() -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Catálogo"
    ws.append(["GTIN", "REFERENCIA", "DESCRIPCION"])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
